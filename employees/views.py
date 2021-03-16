from django.shortcuts import render
from openpyxl import load_workbook
from dateutil.parser import parse
from django.views.generic import ListView, DetailView, UpdateView
from .models import Employees


def home(request):
    columns, employees_file_data = get_data_from_excel()
    employees = []
    for employees_row_data in employees_file_data:
        employee_data = []
        for index, data in enumerate(employees_row_data):
            if index == 2:
                employee_data.append(int(data))
            elif index == 3:
                dt = parse(str(data), dayfirst=True)
                employee_data.append(dt.strftime("%d/%m/%Y"))
            elif index == 5:
                employee_data.append(data.strip("BGN"))
            else:
                employee_data.append(data)
        employees.append(employee_data)

    return render(request, 'home/company.html', context={'employees': employees, 'columns': columns})


def update(request):
    if request.method == 'GET':
        columns, employees_file_data = get_data_from_excel()

        for employees_row_data in employees_file_data:
            emp = Employees()
            for index, data in enumerate(employees_row_data):
                if index == 0:
                    emp.first_name = data
                if index == 1:
                    emp.last_name = data
                if index == 2:
                    emp.mobile = int(data)
                elif index == 3:
                    parse_date = parse(str(data), dayfirst=True)
                    emp.start_date = parse_date
                elif index == 5:
                    emp.salary = data.strip("BGN")
                elif index == 6:
                    emp.emp_id = data
                else:
                    pass
            emp.save()

    context = {
        'update': True
    }

    return render(request, 'home/update.html', context)


def get_data_from_excel():
    wb = load_workbook('./files/Employee table.xlsx')
    sheets = wb.sheetnames
    tes = wb[sheets[0]]
    columns = None
    employees_file_data = []
    columns_count = 0
    for rows in tes.values:
        if columns_count < 1:
            columns = rows[0:7]
        else:
            employees_file_data.append(rows[0:7])
        columns_count += 1

    return columns, employees_file_data


class EmployeeListView(ListView):
    model = Employees
    template_name = 'home/employees.html'
    context_object_name = 'employees'


class EmployeesDetailView(DetailView):
    model = Employees
    template_name = 'home/employee_detail.html'
    context_object_name = 'employees'


class EmployeesUpdateView(UpdateView):
    model = Employees
    fields = ['first_name', 'last_name', 'salary', 'mobile', 'emp_id']
    template_name = 'home/employee_form.html'
    success_url = '/employees'


class EmployeesDeleteView(DetailView):
    model = Employees
    template_name = 'home/employee_confirm_delete.html'
